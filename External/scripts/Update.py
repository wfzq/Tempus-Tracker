"""
    Database Update Script by wfzq (Amuria)
"""

import json
import time
import requests
import datetime
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from pathlib import Path

"""
    Configure Script
"""

BACKUP_CURRENT_FILE = True
CHECK_FOR_NEW_MAPS = True
LOG_TO_SHEET = True

# Paste exact map name in quotes e.g.: jump_beef 
# to continue from that point, otherwise leave empty
CONTINUE_FROM_MAP = ""

# Update only if a map is deleted 
# from the previous version of the database
DELETED_MAPS = 1 

# Automatically remove old maps_json files
# a legacy file is one not currently in use
# MIN = 1
MAX_LEGACY_FILES = 1

# Locations
MAPS_FOLDER = "src/data/maps_db/"
WORKBOOK_URL = "src/data/maps_db/DB_Completions.xlsx"
MAPS_JSON_URL = "src/data/maps_db/maps_merged.json"

# APIs
GET_MAP_LIST = "https://tempus2.xyz/api/v0/maps/detailedList"
GET_MAP_COMP = "https://tempus2.xyz/api/v0/maps/id/{}/zones/typeindex/map/1/records/player/0/3"

MAP_IDS = set()


def get_map_template():
    return {
        "intended_class": None,
        "completion_info": {
            "soldier": None,
            "demoman": None,
            "last_updated": None
        },
        "class_tech": {
            "soldier": [],
            "demoman": []
        }
    }


def get_map_completions(map_id, max_retries=3, base_delay=1.0):
    url = GET_MAP_COMP.format(map_id)
    retries = 0

    while retries < max_retries:
        try:
            response = requests.get(url, timeout=15)
            if response.status_code == 429:
                # Try to get retry-after header, default to exponential backoff
                retry_after = float(response.headers.get('Retry-After', base_delay * (2 ** retries)))
                print(f"Rate limited at: {map_id}. Waiting {retry_after} seconds")
                time.sleep(retry_after)
                retries += 1
                continue
            
            
            response.raise_for_status()
            data = response.json()
            completion_info = data.get('completion_info', {})
            completion_info['last_updated'] = datetime.datetime.now().isoformat()

            return {
                "name": map_id,
                "completion_info": completion_info
            }

        except requests.RequestException as e:
            if hasattr(e, 'response') and e.response is not None:
                
                if e.response.status_code == 429:
                    retry_after = float(e.response.headers.get('Retry-After', base_delay * (2 ** retries)))
                    time.sleep(retry_after)
                    print(f"Rate limited when fetching map id: {map_id}. Retrying after {retry_after} seconds")
                
                elif e.response.status_code == 404:
                    print(f"Map id: {map_id} not found")
                    break
                else:
                    print(f"Failed to fetch data for map id: {map_id}. \
                        Status code: {e.response.status_code}")
            else:
                print(f"Network error when fetching map id: {map_id}. Error: {str(e)}")
            
            retries += 1
            if retries < max_retries:
                # Exponential backoff
                delay = base_delay * (2 ** retries)
                print(f"Retrying {retries}/{max_retries} after {delay} seconds...")
                time.sleep(delay)
        except json.JSONDecodeError:
            print("Error parsing API response as JSON")
    
    return None


def fill_map_info(m):
    # Terrible, I know
    O_tech = ["Mixed", "Pogo", "Gimmick", "Phase", "Strafe", "Jank", "Jurf", "Edgebug", "Wallbug", "Texturebug", "Hole", "Lim Ammo", "Technical", "Bounce", "Buttons"]
    S_tech = ["Wallpogo", "Wallshot", "Sync", "Prefire", "Rand Bounce", "Airstrike", "Speedshot", "Ctap"]
    D_tech = ["3pre", "Hardpogo", "Airpogo", "Downair", "Vert"]

    # Set intended class
    print(f"\nIs \033[92m{m['name']}\033[0m soldier or demo intended?")
    print("s|3 - Soldier\nd|4 - Demoman\nEnter - Leave blank")
    while True:
        intended_class = input(">> ").lower()

        if intended_class == 's' or intended_class == '3':
            intended_class = 3
            break
        elif intended_class == 'd' or intended_class == '4':
            intended_class = 4
            break
        elif intended_class == '':
            intended_class = None
            break
    m["intended_class"] = intended_class

    # Set Soldier tech
    print("")
    print("List solly tech separated by commas such as: mixed, pogo, ctap or perss Enter to leave blank")
    s_input = input(">> ")
    s_array = [item.strip().capitalize() for item in s_input.split(",")]

    for i, item in enumerate(s_array):
        if s_input == '':
            break
        
        if item not in O_tech and item not in S_tech:
            resolved = False
            while not resolved:
                print(f"{item} is not an existing \033[94msoldier\033[0m tech, select an action: ")
                print("1 - Add element")
                print("2 - Append element")
                print("3 - Remove element")

                ans = input(">> ").lower()
                if ans == '1':
                    resolved = True

                elif ans == '2':
                    s_array[i] = input("New Element: ").capitalize()
                    item = s_array[i]
                    if s_array[i] in O_tech or s_array[i] in S_tech:
                        resolved = True

                elif ans == '3': 
                    s_array[i] = ""
                    resolved = True

                else:
                    continue
        
        if s_array[i] != "":
            m["class_tech"]["soldier"].append(s_array[i])


    # Set Demoman tech
    print("")
    print("List demo tech separated by commas such as: mixed, pogo, 3pre or perss Enter to leave blank")
    s_input = input(">> ")
    s_array = [item.strip().capitalize() for item in s_input.split(",")]

    for i, item in enumerate(s_array):
        if s_input == '':
            break
        if item not in O_tech and item not in D_tech:
            resolved = False
            while not resolved:
                print(f"{item} is not an existing \033[1;91mdemoman\033[0m tech, select an action: ")
                print("1 - Add element")
                print("2 - Append element")
                print("3 - Remove element")

                ans = input(">> ").lower()
                if ans == '1':
                    resolved = True

                elif ans == '2':
                    s_array[i] = input("New Element: ").capitalize()
                    item = s_array[i]
                    if s_array[i] in O_tech or s_array[i] in S_tech:
                        resolved = True

                elif ans == '3': 
                    s_array[i] = ""
                    resolved = True

                else:
                    continue
        
        if s_array[i] != "":
            m["class_tech"]["demoman"].append(s_array[i])


def update_maps(maps_json):
    print("Checking for new maps...")
    # Fetch maps from API
    try:
        response = requests.get(GET_MAP_LIST)
        response.raise_for_status()
        updated_map_list = response.json()
    except requests.RequestException as e:
        print(f"API request failed: {e}")
        exit(1)
    except json.JSONDecodeError:
        print("Error parsing API response as JSON")
        exit(1)
    
    # Create 'id' dictionary
    local_maps_by_id = {item.get('id'): item for item in maps_json if 'id' in item}
    updated_elements = []

    # Process each item from the API response
    for updated_map in updated_map_list:
        api_id = updated_map['id']
        MAP_IDS.add(api_id)
        new_map = False

        # Use existing map or create a new one
        if api_id in local_maps_by_id:
            current_map = local_maps_by_id[api_id]
        else:
            print(f"Added new map: {updated_map.get('name')}")
            current_map = updated_map.copy()
            new_map = True

        # Check for missing fields
        for key, value in get_map_template().items():
            if key not in current_map:
                print(f"Added missing: {key} in {current_map.get('name')}")
                current_map[key] = value
                
            
            # Check for missing subkeys
            elif isinstance(value, dict) and isinstance(current_map[key], dict):
                for subkey, subvalue in value.items():
                    if subkey not in current_map[key]:
                        current_map[key][subkey] = subvalue
                        print(f"Added missing: {subkey} in {key}, for map {current_map.get('name')}")
    
        if new_map:
            fill_map_info(current_map)
            new_map = False

        # Merge with API data
        merged_map = {**current_map, **updated_map}
        updated_elements.append(merged_map)

    # Remove local items that weren't in the API response
    for local_id, current_map in local_maps_by_id.items():
        if local_id not in MAP_IDS:
            maps_json.remove(current_map)
            print(f"Map: {current_map.get('name')} removed from map pool")
            
    print("Maps up to date! Saving..\n")
    save_mapsmerged(updated_elements)

    return updated_elements

def update_map_completions(maps_json, start_from_map_name=None):
    start_from_map = bool(start_from_map_name)

    for i, current_map in enumerate(maps_json):
        map_id = current_map['id']

        if start_from_map:
            if current_map['name'] == start_from_map_name:
                print(f"Starting from map: {map_id}")
                start_from_map = None
            else:
                continue

        result = get_map_completions(map_id)
        
        if result:
            print(f"Updated completions for: {maps_json[i]['name']} ({i + 1}/{len(maps_json)})")
            maps_json[i]['completion_info'] = result['completion_info']

            # Save every 10 maps and after completing the last one
            if (i + 1) % 10 == 0 or i == len(maps_json) - 1:
                with open(MAPS_JSON_URL, 'w') as f:
                    json.dump(maps_json, f, indent=2)
                print(f"Progress saved until: {maps_json[i]['name']}")

        
        else:
            print(f"Failed to get completions for map id: {map_id}")
    
    if start_from_map != None:
        print(f"Map: {start_from_map_name} not found in file, check spelling or capitalization")

    return maps_json

def clean_legacy_files():
    if MAX_LEGACY_FILES < 1: 
        print("ERR: You need at least 1 backup file")
        exit(1)

    json_files = [f for f in Path(MAPS_FOLDER).glob("maps_merged*.json")]
    json_files.sort(key=lambda f: f.stat().st_mtime)

    to_delete = len(json_files) - (1 + MAX_LEGACY_FILES)
    
    if to_delete >= 1:
        for i in range(to_delete):
            try:
                json_files[i].unlink()
                print(f"Deleted: {json_files[i].name}")
            except Exception as e:
                print(f"Failed to delete {json_files[i].name}: {e}")
 

def open_maps_merged():
    try:
        with open(MAPS_JSON_URL, 'r') as f:
            return json.load(f)

    except FileNotFoundError:
        print(f"Local file not found: {MAPS_JSON_URL}")
        exit(1)
    except json.JSONDecodeError:
        print(f"Error parsing local JSON file: {MAPS_JSON_URL}")
        exit(1)


def save_mapsmerged(maps_json):
    try:
        with open(MAPS_JSON_URL, 'w') as f:
            json.dump(maps_json, f, indent=2)
        print(f"Successfully updated maps_merged.json")

    except Exception as e:
        print(f"Error saving updated data: {e}")
        exit(1)


def make_backup(file_path):
    print("Creating backup...")

    if os.path.exists(file_path):
        directory, filename = os.path.split(file_path)
        filename_without_ext, ext = os.path.splitext(filename)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file_path = os.path.join(
            directory, f"{filename_without_ext}_{timestamp}.json")

        shutil.copyfile(file_path, backup_file_path)
        print(f"Backup successfull!\n")
    else:
        print(f"File not found: {file_path}, exiting...")
        exit(1)

def log_completions_to_sheet(maps_json):
    wb = load_workbook(WORKBOOK_URL)
    ws = wb.active
    if ws is None:
        print("Failed to open active sheet")
        exit(1)

    new_line = []

    # Sort maps by ID
    maps_json.sort(key=lambda f: f["id"])
    # Get sheet maps list (ordered by ID)
    sheet_names = [cell.value for cell in ws[1][1::2]]

    # Get timestmap
    for current in maps_json:
        timestamp = current.get("completion_info").get("last_updated")
        if timestamp:
            timestamp = timestamp.split("T")[0]
            break
        else:
            print(f"Timestamp unavailable for map: {current.name}")
    
    if timestamp:
        new_line.append(timestamp)
    else:
        print("ERR: No timestamps found")
        exit(1)

    # Write new completions
    db_idx = 0
    db_deleted = 0
    db_deleted_names = []
    for current in maps_json:
        name = current["name"]
        sc = current["completion_info"]["soldier"]
        dc = current["completion_info"]["demoman"]

        if not (name or sc or dc):
            print(f"ERR: Missing data on: {name}")
            exit(1)

        # Database has elements and the current element doesn't match
        while db_idx < len(sheet_names) and name != sheet_names[db_idx]:
            print(f"\n>>> {name} != {sheet_names[db_idx]}")
            if input("WARN: Name mismatch, is this the same map? (y/N): ").lower() == 'y':
                sheet_names[db_idx] = name
                ws.cell(row=1, column=2 + db_idx*2, value=name)
            else:
                new_line.append("-")
                new_line.append("-")

                db_deleted_names.append(sheet_names[db_idx])
                db_deleted += 1
                db_idx += 1

        new_line.append(sc)
        new_line.append(dc)

        # Database has NO elements, add one
        if db_idx >= len(sheet_names):
            cell = ws.cell(row=1, column=1 + db_idx*2 + 1, value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(
                f"{get_column_letter(2 + db_idx*2)}1:{get_column_letter(2 + db_idx*2 + 1)}1"
            )

        db_idx += 1

    print(f"{db_deleted} deleted maps")
    if db_deleted > 0:
        print(f"First deleted map: {db_deleted_names[0]}")
    
    # FATAL ERROR in case of map name change
    if db_deleted > DELETED_MAPS:
        print("ERR: Are there really that many deleted maps?")
        print("ERR: This is probably due to a map name change!")
        exit(1)

    ws.append(new_line)

    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(WORKBOOK_URL)
    return


if __name__ == "__main__":
    if BACKUP_CURRENT_FILE: 
        make_backup(MAPS_JSON_URL)

    maps_json = open_maps_merged()

    if CHECK_FOR_NEW_MAPS:
        maps_json = update_maps(maps_json)
   
    update_map_completions(maps_json, CONTINUE_FROM_MAP)
    
    clean_legacy_files()

    if LOG_TO_SHEET:
        log_completions_to_sheet(maps_json)